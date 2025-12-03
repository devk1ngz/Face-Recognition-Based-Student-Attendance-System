import os
import cv2
import pandas as pd
import numpy as np
import math
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel, 
                             QPushButton, QFrame, QTableWidget, QTableWidgetItem, 
                             QHeaderView, QDateEdit, QLineEdit, QGraphicsDropShadowEffect,
                             QComboBox, QListView, QCalendarWidget, QToolButton,
                             QFileDialog, QMessageBox, QDialog, QGridLayout, QAbstractItemView)
from PyQt5.QtCore import Qt, QDate, QLocale, QTimer
from PyQt5.QtGui import QColor, QPixmap, QImage, QPainter, QPainterPath, QPen, QFont
from app.config import IMG_DIR, STYLES_DIR
from PyQt5.QtCore import QRect
from app.controllers.teacher_controller import (get_teacher_courses, get_attendance_list, 
                                                update_attendance, add_student, get_teacher_detail, get_teaching_stats,
                                                create_or_get_course, import_student_from_excel)
from PyQt5.QtWidgets import QButtonGroup
from app.core.camera_dialog import CameraDialog


class TeacherProfileDialog(QDialog):
    def __init__(self, teacher_id, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Hồ sơ giảng viên")
        self.setFixedSize(700, 700)
        self.setStyleSheet("background-color: white; font-family: 'Segoe UI';")
        self.t_info = get_teacher_detail(teacher_id)
        self.courses_data = get_teaching_stats(teacher_id)
        self.init_ui()

    def init_ui(self):
        layout = QVBoxLayout()
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(25)
        self.setLayout(layout)
        header_layout = QVBoxLayout()
        header_layout.setSpacing(15)
        
        self.ava = QLabel()
        self.ava.setFixedSize(140, 140)
        self.ava.setAlignment(Qt.AlignCenter)
        self.load_avatar_circle()
        
        lbl_name = QLabel(self.t_info.fullname.upper() if self.t_info else "UNKNOWN")
        lbl_name.setStyleSheet("font-size: 24px; font-weight: 900; color: #2c3e50; letter-spacing: 1px;")
        lbl_name.setAlignment(Qt.AlignCenter)
        
        lbl_code = QLabel(f"Mã GV: {self.t_info.user_code}" if self.t_info else "---")
        lbl_code.setStyleSheet("font-size: 15px; color: #7f8c8d; font-weight: bold; background: #f1f2f6; padding: 5px 15px; border-radius: 15px;")
        lbl_code.setAlignment(Qt.AlignCenter)
        
        header_layout.addWidget(self.ava, alignment=Qt.AlignCenter)
        header_layout.addWidget(lbl_name)
        header_layout.addWidget(lbl_code, alignment=Qt.AlignCenter)
        layout.addLayout(header_layout)
        line = QFrame(); line.setFrameShape(QFrame.HLine); line.setStyleSheet("color: #ecf0f1;")
        layout.addWidget(line)
        info_container = QFrame(); info_container.setStyleSheet("background-color: white;")
        grid = QGridLayout(); grid.setContentsMargins(20, 0, 20, 0); grid.setHorizontalSpacing(30); grid.setVerticalSpacing(15)
        
        style_lbl = "color: #95a5a6; font-size: 14px; font-weight: 600;"
        style_val = "color: #2c3e50; font-size: 16px; font-weight: bold;"

        phone = self.t_info.phone_number if (self.t_info and self.t_info.phone_number) else "Chưa cập nhật"
        if len(phone) >= 10 and phone.isdigit(): phone = f"{phone[:4]}.{phone[4:7]}.{phone[7:]}"

        email = self.t_info.email if (self.t_info and self.t_info.email) else "Chưa cập nhật"

        grid.addWidget(QLabel("Số điện thoại:"), 0, 0); l_p = QLabel(phone); l_p.setStyleSheet(style_val); grid.addWidget(l_p, 0, 1)
        grid.addWidget(QLabel("Email:"), 1, 0); l_e = QLabel(email); l_e.setStyleSheet(style_val); grid.addWidget(l_e, 1, 1)
        
        for i in range(grid.rowCount()):
            item = grid.itemAtPosition(i, 0).widget()
            if item: item.setStyleSheet(style_lbl)
            
        info_container.setLayout(grid)
        layout.addWidget(info_container)

        layout.addSpacing(10)
        lbl_table = QLabel("CÁC HỌC PHẦN ĐANG GIẢNG DẠY")
        lbl_table.setStyleSheet("font-size: 14px; font-weight: 800; color: #34495e; text-transform: uppercase; border-left: 4px solid #e67e22; padding-left: 10px;")
        layout.addWidget(lbl_table)
        
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["MÃ HP", "TÊN HỌC PHẦN", "SĨ SỐ"])
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.table.setShowGrid(False)
        self.table.setFocusPolicy(Qt.NoFocus)
        
        self.table.setStyleSheet("""
            QTableWidget { border: 1px solid #dfe6e9; border-radius: 8px; font-size: 14px; }
            QHeaderView::section { background-color: #f8f9fa; border: none; padding: 10px; font-weight: bold; color: #7f8c8d; }
            QTableWidget::item { padding: 8px; border-bottom: 1px solid #f1f2f6; }
        """)
        
        h = self.table.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.Fixed); self.table.setColumnWidth(0, 120)
        h.setSectionResizeMode(1, QHeaderView.Stretch)
        h.setSectionResizeMode(2, QHeaderView.Fixed); self.table.setColumnWidth(2, 100)

        self.table.setRowCount(len(self.courses_data))
        for r, data in enumerate(self.courses_data):
            it_code = QTableWidgetItem(data['code']); it_code.setTextAlignment(Qt.AlignCenter); self.table.setItem(r, 0, it_code)
            it_name = QTableWidgetItem(data['name']); it_name.setTextAlignment(Qt.AlignCenter); self.table.setItem(r, 1, it_name)
            it_count = QTableWidgetItem(str(data['count'])); it_count.setTextAlignment(Qt.AlignCenter); self.table.setItem(r, 2, it_count)
            
        layout.addWidget(self.table)
        
        btn_close = QPushButton("Đóng hồ sơ")
        btn_close.setCursor(Qt.PointingHandCursor)
        btn_close.setStyleSheet("QPushButton { background-color: #34495e; color: white; padding: 12px; border-radius: 8px; font-weight: bold; border: none; } QPushButton:hover { background-color: #2c3e50; }")
        btn_close.clicked.connect(self.close)
        layout.addWidget(btn_close)

    def load_avatar_circle(self):
        size = 140
        final_path = None
        code = self.t_info.user_code if self.t_info else ""

        paths = [
            os.path.join(IMG_DIR, "teacher", f"{code}.jpg"),
            os.path.join(IMG_DIR, "teacher", f"{code}.png"),
            os.path.join(IMG_DIR, "teacher", f"{code.lower()}.jpg")
        ]
        for p in paths:
            if os.path.exists(p): final_path = p; break
            
        if not final_path: final_path = os.path.join(IMG_DIR, "teacher_avatar.png")

        rounded = QPixmap(size, size); rounded.fill(Qt.transparent)
        painter = QPainter(rounded)
        painter.setRenderHint(QPainter.Antialiasing, True); painter.setRenderHint(QPainter.SmoothPixmapTransform, True)
        path = QPainterPath(); path.addEllipse(0, 0, size, size); painter.setClipPath(path)
        
        if os.path.exists(final_path):
             painter.drawPixmap(0, 0, QPixmap(final_path).scaled(size, size, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation))
        else:
             colors = ["#2980b9", "#8e44ad", "#27ae60", "#d35400", "#c0392b"]
             idx = len(code) % len(colors)
             painter.setBrush(QColor(colors[idx])); painter.setPen(Qt.NoPen); painter.drawRect(0, 0, size, size)
             painter.setPen(QColor("white")); font = QFont("Segoe UI", 40, QFont.Bold); painter.setFont(font)
             painter.drawText(QRect(0, 0, size, size), Qt.AlignCenter, code[:2])

        painter.setClipping(False); pen = QPen(QColor("#e67e22")); pen.setWidth(4); painter.setPen(pen); painter.setBrush(Qt.NoBrush)
        painter.drawEllipse(2, 2, size-4, size-4); painter.end()
        self.ava.setPixmap(rounded)

# =================================================================================
# CLASS 1: BADGE TRẠNG THÁI 
# =================================================================================
class AttendanceStatusBox(QComboBox):
    def __init__(self, current_status="Chưa điểm danh"):
        super().__init__()
        self.setCursor(Qt.PointingHandCursor)
        self.addItems(["Có mặt", "Vắng", "Chưa điểm danh"])
        self.setCurrentText(current_status)
        self.setView(QListView(self)) 
        self.update_style()
        self.currentTextChanged.connect(self.update_style)

    def update_style(self):
        text = self.currentText()
        if text == "Có mặt": 
            bg_color = "#e6fffa"; text_color = "#00b894"; border_color = "#00b894"
        elif text == "Vắng": 
            bg_color = "#fff5f5"; text_color = "#d63031"; border_color = "#d63031"
        else: 
            bg_color = "#f1f2f6"; text_color = "#747d8c"; border_color = "#ced6e0"
            
        self.setStyleSheet(f"""
            QComboBox {{
                background-color: {bg_color}; 
                color: {text_color}; 
                border: 1px solid {border_color};
                border-radius: 12px;
                padding: 4px 10px; 
                font-weight: bold;
                font-family: "Segoe UI"; font-size: 13px; 
                min-width: 110px; margin: 5px 5px;
            }}
            QComboBox::drop-down {{ border: none; width: 0px; }}
            QComboBox::down-arrow {{ image: none; }}
            QComboBox QAbstractItemView {{
                background-color: white; color: #333; 
                selection-background-color: #0984e3; selection-color: white;
                outline: none; padding: 5px; border: 1px solid #dfe6e9;
            }}
        """)

# =================================================================================
# CLASS 2: DIALOG THÊM SINH VIÊN (CÓ CAMERA)
# =================================================================================
class AddStudentDialog(QDialog):
    def __init__(self, course_id, parent=None):
        super().__init__(parent)
        self.course_id = course_id
        self.captured_frame = None
        self.cap = None
        
        self.setWindowTitle("Thêm Sinh Viên Mới")
        self.setFixedSize(950, 550)
        self.setWindowModality(Qt.ApplicationModal)
        self.setStyleSheet("background-color: white; font-family: 'Segoe UI'; font-size: 14px;")
        
        self.init_ui()
        self.start_camera()

    def init_ui(self):
        layout = QHBoxLayout()
        self.setLayout(layout)
        left_layout = QVBoxLayout()
        left_layout.setContentsMargins(20, 20, 20, 20)
        left_layout.setSpacing(15)
        
        lbl_title = QLabel("THÔNG TIN SINH VIÊN")
        lbl_title.setStyleSheet("color: #2c3e50; font-weight: 800; font-size: 22px; margin-bottom: 10px;")
        left_layout.addWidget(lbl_title)

        self.txt_mssv = QLineEdit(); self.txt_mssv.setPlaceholderText("Nhập Mã số SV...")
        self.txt_name = QLineEdit(); self.txt_name.setPlaceholderText("Nhập Họ và Tên...")
        self.txt_class = QLineEdit(); self.txt_class.setPlaceholderText("Nhập Lớp hành chính...")
        self.date_dob = QDateEdit(); self.date_dob.setCalendarPopup(True); self.date_dob.setDisplayFormat("dd-MM-yyyy")
        self.date_dob.setDate(QDate(2004, 1, 1)) 
        
        inp_style = """
            QLineEdit, QDateEdit { 
                border: 1px solid #dcdfe6; border-radius: 8px; 
                padding: 10px; background: #f8f9fa; font-size: 15px;
            }
            QLineEdit:focus, QDateEdit:focus { border: 1px solid #3498db; background: white; }
        """
        for w in [self.txt_mssv, self.txt_name, self.txt_class, self.date_dob]:
            w.setStyleSheet(inp_style); w.setFixedHeight(45)

        left_layout.addWidget(QLabel("Mã số sinh viên:"))
        left_layout.addWidget(self.txt_mssv)
        left_layout.addWidget(QLabel("Họ và Tên:"))
        left_layout.addWidget(self.txt_name)
        left_layout.addWidget(QLabel("Ngày sinh:"))
        left_layout.addWidget(self.date_dob)
        left_layout.addWidget(QLabel("Lớp:"))
        left_layout.addWidget(self.txt_class)
        left_layout.addStretch()
        
        self.btn_save = QPushButton(" LƯU THÔNG TIN")
        self.btn_save.setCursor(Qt.PointingHandCursor)
        self.btn_save.setStyleSheet("""
            QPushButton { background-color: #27ae60; color: white; border-radius: 8px; font-weight: bold; padding: 12px; border: none; }
            QPushButton:hover { background-color: #219150; }
        """)
        self.btn_save.clicked.connect(self.save_student)
        left_layout.addWidget(self.btn_save)

        right_layout = QVBoxLayout()
        right_layout.setContentsMargins(10, 20, 20, 20)
        
        self.lbl_camera = QLabel("Đang tải Camera...")
        self.lbl_camera.setFixedSize(450, 350)
        self.lbl_camera.setStyleSheet("background-color: #2d3436; border-radius: 12px; color: white;")
        self.lbl_camera.setAlignment(Qt.AlignCenter)
        
        self.btn_capture = QPushButton(" CHỤP ẢNH")
        self.btn_capture.setCursor(Qt.PointingHandCursor)
        self.btn_capture.setStyleSheet("""
            QPushButton { background-color: #3498db; color: white; border-radius: 8px; font-weight: bold; padding: 12px; border: none; }
            QPushButton:hover { background-color: #2980b9; }
        """)
        self.btn_capture.clicked.connect(self.capture_image)
        
        right_layout.addWidget(self.lbl_camera)
        right_layout.addWidget(self.btn_capture)
        right_layout.addStretch()
        
        layout.addLayout(left_layout, 4)
        layout.addLayout(right_layout, 6)

    def start_camera(self):
        self.cap = cv2.VideoCapture(0)
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_frame)
        self.timer.start(30)

    def update_frame(self):
        if self.cap and self.cap.isOpened():
            ret, frame = self.cap.read()
            if ret:
                frame = cv2.flip(frame, 1)
                self.current_frame_cv = frame
                rgb_img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                h, w, ch = rgb_img.shape
                bytes_per_line = ch * w
                qt_img = QImage(rgb_img.data, w, h, bytes_per_line, QImage.Format_RGB888)
                self.lbl_camera.setPixmap(QPixmap.fromImage(qt_img).scaled(450, 350, Qt.KeepAspectRatio, Qt.SmoothTransformation))

    def capture_image(self):
        if hasattr(self, 'current_frame_cv') and self.current_frame_cv is not None:
            self.captured_frame = self.current_frame_cv.copy()
            self.timer.stop() 
            self.btn_capture.setText(" ĐÃ CHỤP (Nhấn để chụp lại)")
            self.btn_capture.setStyleSheet("background-color: #e67e22; color: white; border-radius: 8px; font-weight: bold; padding: 12px; border: none;")
            self.btn_capture.clicked.disconnect()
            self.btn_capture.clicked.connect(self.retake_image)

    def retake_image(self):
        self.captured_frame = None
        self.timer.start(30)
        self.btn_capture.setText(" CHỤP ẢNH")
        self.btn_capture.setStyleSheet("background-color: #3498db; color: white; border-radius: 8px; font-weight: bold; padding: 12px; border: none;")
        self.btn_capture.clicked.disconnect()
        self.btn_capture.clicked.connect(self.capture_image)

    def save_student(self):
        mssv = self.txt_mssv.text().strip()
        name = self.txt_name.text().strip()
        dob = self.date_dob.date().toPyDate() 
        class_name = self.txt_class.text().strip()
        
        if not mssv or not name:
            QMessageBox.warning(self, "Thiếu thông tin", "Vui lòng nhập MSSV và Tên!")
            return
            
        if self.captured_frame is None:
            QMessageBox.warning(self, "Thiếu ảnh", "Vui lòng chụp ảnh khuôn mặt!")
            return
            
        success, msg = add_student(mssv, name, dob, class_name, self.course_id, self.captured_frame)
        
        if success:
            QMessageBox.information(self, "Thành công", msg)
            self.close()
        else:
            QMessageBox.critical(self, "Lỗi", msg)

    def closeEvent(self, event):
        if self.cap: self.cap.release()
        event.accept()


# =================================================================================
# CLASS 3: CỬA SỔ CHÍNH GIÁO VIÊN
# =================================================================================
class TeacherWindow(QWidget):
    def __init__(self, teacher_id=None):
        super().__init__()
        self.teacher_id = teacher_id
        self.current_course_id = None
        self.all_table_data = [] 
        
        self.setWindowTitle("Hệ Thống Quản Lý Điểm Danh - Giáo Viên")
        self.setMinimumSize(1300, 800)
        self.initUI()
        self.load_style()
        self.load_sidebar_data()

    def load_style(self):
        path = os.path.join(STYLES_DIR, "teacher_style.qss")
        if os.path.exists(path):
            with open(path, "r", encoding="utf-8") as f: self.setStyleSheet(f.read())

    def initUI(self):
        main_layout = QHBoxLayout()
        main_layout.setContentsMargins(0, 0, 0, 0)
        main_layout.setSpacing(0)
        self.setLayout(main_layout)

        self.sidebar = QFrame()
        self.sidebar.setObjectName("sidebar")
        self.sidebar.setFixedWidth(330)
        self.sidebar_layout = QVBoxLayout()
        self.sidebar_layout.setContentsMargins(0, 30, 0, 20)
        self.sidebar_layout.setSpacing(10)
        self.sidebar.setLayout(self.sidebar_layout)

        self.setup_profile_section(self.sidebar_layout)
        self.sidebar_layout.addSpacing(30)
        
        lbl_menu = QLabel("DANH SÁCH LỚP HỌC")
        lbl_menu.setObjectName("sidebar_title")
        self.sidebar_layout.addWidget(lbl_menu)

        self.courses_container = QVBoxLayout()
        self.courses_container.setSpacing(2)
        self.sidebar_layout.addLayout(self.courses_container)
        self.sidebar_layout.addStretch()

        self.btn_logout = QPushButton("Đăng xuất")
        self.btn_logout.setObjectName("btn_logout")
        self.btn_logout.setCursor(Qt.PointingHandCursor)
        self.sidebar_layout.addWidget(self.btn_logout)

        self.content_area = QFrame()
        self.content_area.setObjectName("content_area")
        content_layout = QVBoxLayout()
        content_layout.setContentsMargins(40, 40, 40, 40)
        content_layout.setSpacing(25)
        self.content_area.setLayout(content_layout)

        header_layout = QHBoxLayout()
        header_layout.setContentsMargins(0, 0, 0, 10)
        
        self.logo_small = QLabel()
        self.logo_small.setFixedSize(80, 80)
        logo_path = os.path.join(IMG_DIR, "logo_humg.jpg")
        if os.path.exists(logo_path):
            self.logo_small.setPixmap(QPixmap(logo_path).scaled(80, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation))

        separator = QFrame()
        separator.setFrameShape(QFrame.VLine)
        separator.setFrameShadow(QFrame.Sunken)
        separator.setFixedHeight(60) 
        separator.setStyleSheet("color: #bdc3c7; border: 1px solid #bdc3c7;")
        
        title_layout = QVBoxLayout()
        title_layout.setSpacing(2)
        title_layout.setContentsMargins(0, 5, 0, 5)
        lbl_school = QLabel("TRƯỜNG ĐẠI HỌC MỎ - ĐỊA CHẤT")
        lbl_school.setObjectName("school_name")
        self.lbl_class = QLabel("Chọn lớp học...")
        self.lbl_class.setObjectName("header_title")
        self.lbl_sub_info = QLabel("") 
        self.lbl_sub_info.setStyleSheet("color: #7f8c8d; font-size: 16px; font-weight: 600; margin-top: 2px;")
        title_layout.addWidget(lbl_school)
        title_layout.addWidget(self.lbl_class)
        title_layout.addWidget(self.lbl_sub_info)
        
        header_layout.addWidget(self.logo_small)
        header_layout.addSpacing(15) 
        header_layout.addWidget(separator) 
        header_layout.addSpacing(15)
        header_layout.addLayout(title_layout)
        header_layout.addStretch()
        
        date_layout = QHBoxLayout()
        lbl_date = QLabel("Ngày điểm danh:")
        lbl_date.setStyleSheet("font-weight: 600; color: #7f8c8d; font-size: 14px;")
        self.date_edit = QDateEdit()
        self.date_edit.setCalendarPopup(True)
        self.date_edit.setDate(QDate.currentDate())
        self.date_edit.setFixedWidth(130); self.date_edit.setFixedHeight(35)
        self.date_edit.setStyleSheet("""
            QDateEdit { 
                border: 1px solid #dfe6e9; border-radius: 8px; 
                padding: 5px 10px; background: white; color: #2c3e50; font-weight: bold;
            }
            QDateEdit::drop-down { border: none; }
            QDateEdit::down-arrow { image: none; } 
        """)
        self.date_edit.dateChanged.connect(self.refresh_table)
        date_layout.addWidget(lbl_date)
        date_layout.addWidget(self.date_edit)
        
        header_layout.addLayout(date_layout)

        stats_layout = QHBoxLayout()
        stats_layout.setSpacing(20)
        self.card_total = self.create_stat_card("Tổng Sĩ Số", "0", "#3498db")   
        self.card_present = self.create_stat_card("Hiện Diện", "0", "#27ae60")  
        self.card_absent = self.create_stat_card("Vắng Mặt", "0", "#e74c3c")    
        stats_layout.addWidget(self.card_total)
        stats_layout.addWidget(self.card_present)
        stats_layout.addWidget(self.card_absent)

        toolbar_layout = QHBoxLayout()
        
        self.btn_cam = QPushButton(" BẮT ĐẦU ĐIỂM DANH")
        self.btn_cam.setObjectName("btn_primary")
        self.btn_cam.setCursor(Qt.PointingHandCursor)
        self.btn_cam.clicked.connect(self.open_camera)

        self.btn_excel = QPushButton(" XUẤT EXCEL")
        self.btn_excel.setObjectName("btn_success")
        self.btn_excel.setCursor(Qt.PointingHandCursor)
        self.btn_excel.clicked.connect(self.export_excel)

        self.btn_add = QPushButton(" THÊM SINH VIÊN")
        self.btn_add.setObjectName("btn_purple")
        self.btn_add.setCursor(Qt.PointingHandCursor)
        self.btn_add.clicked.connect(self.open_add_student)

        self.btn_import = QPushButton(" NHẬP FILE LỚP")
        self.btn_import.setObjectName("btn_import_class") 
        self.btn_import.setCursor(Qt.PointingHandCursor)
        self.btn_import.clicked.connect(self.handle_import_excel)

        toolbar_layout.addWidget(self.btn_cam)
        toolbar_layout.addSpacing(10)
        toolbar_layout.addWidget(self.btn_excel)
        toolbar_layout.addSpacing(10)
        toolbar_layout.addWidget(self.btn_import)
        toolbar_layout.addSpacing(10)   
        toolbar_layout.addWidget(self.btn_add)
        
        toolbar_layout.addStretch() 
        self.txt_search = QLineEdit()
        self.txt_search.setObjectName("search_box")
        self.txt_search.setPlaceholderText("Tìm kiếm tên hoặc MSSV...")
        self.txt_search.setFixedWidth(250)
        self.txt_search.textChanged.connect(self.filter_rows)

        self.cbb_filter = QComboBox()
        self.cbb_filter.setObjectName("filter_box")
        self.cbb_filter.addItems(["Tất cả", "Có mặt", "Vắng", "Chưa điểm danh"])
        self.cbb_filter.currentTextChanged.connect(self.filter_rows)

        toolbar_layout.addWidget(self.txt_search)
        toolbar_layout.addSpacing(10)
        toolbar_layout.addWidget(self.cbb_filter)

        self.table_card = QFrame()
        self.table_card.setObjectName("table_card")
        t_shadow = QGraphicsDropShadowEffect()
        t_shadow.setBlurRadius(20); t_shadow.setColor(QColor(0,0,0,20)); t_shadow.setYOffset(5)
        self.table_card.setGraphicsEffect(t_shadow)
        
        t_layout = QVBoxLayout()
        t_layout.setContentsMargins(0,0,0,0)
        self.table_card.setLayout(t_layout)

        self.table = QTableWidget()
        self.table.setColumnCount(9) 
        self.table.setHorizontalHeaderLabels(["STT", "MSSV", "Họ đệm", "Tên", "Ngày sinh", "Lớp", "Trạng thái", "Thời gian", "Ghi chú"])
        
        header = self.table.horizontalHeader()

        header.setSectionResizeMode(0, QHeaderView.Fixed); self.table.setColumnWidth(0, 70)   # STT
        header.setSectionResizeMode(1, QHeaderView.Fixed); self.table.setColumnWidth(1, 150)  # MSSV 
        header.setSectionResizeMode(2, QHeaderView.Stretch)                                   # Họ đệm 
        header.setSectionResizeMode(3, QHeaderView.Fixed); self.table.setColumnWidth(3, 150)  # Tên 
        header.setSectionResizeMode(4, QHeaderView.Fixed); self.table.setColumnWidth(4, 150)  # Ngày sinh
        header.setSectionResizeMode(5, QHeaderView.Fixed); self.table.setColumnWidth(5, 150)  # Lớp
        header.setSectionResizeMode(6, QHeaderView.Fixed); self.table.setColumnWidth(6, 170)  # Trạng thái
        header.setSectionResizeMode(7, QHeaderView.ResizeToContents)                          # Thời gian
        header.setSectionResizeMode(8, QHeaderView.Stretch)                                   # Ghi chú
        
        self.table.verticalHeader().setVisible(False)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        self.table.setShowGrid(False)
        t_layout.addWidget(self.table)

        content_layout.addLayout(header_layout)
        content_layout.addLayout(stats_layout)
        content_layout.addSpacing(10)
        content_layout.addLayout(toolbar_layout)
        content_layout.addWidget(self.table_card)

        main_layout.addWidget(self.sidebar)
        main_layout.addWidget(self.content_area)

    # --- LOGIC FUNCTIONS ---

    def setup_profile_section(self, layout):
        widget = QWidget()
        l = QVBoxLayout(); l.setContentsMargins(20,10,20,10); l.setSpacing(8); widget.setLayout(l)

        self.avatar = QLabel(); self.avatar.setFixedSize(80,80); self.avatar.setAlignment(Qt.AlignCenter); self.avatar.setText("GV")
        self.avatar.setStyleSheet("background-color: #ecf0f1; border-radius: 40px; color: #34495e; font-weight: bold; font-size: 23px;")

        self.lbl_tname = QLabel("..."); self.lbl_tname.setObjectName("profile_name"); self.lbl_tname.setAlignment(Qt.AlignCenter); self.lbl_tname.setWordWrap(True)
        self.lbl_tid = QLabel("..."); self.lbl_tid.setObjectName("profile_sub"); self.lbl_tid.setAlignment(Qt.AlignCenter)

        self.btn_profile_detail = QPushButton("Xem hồ sơ")
        self.btn_profile_detail.setObjectName("btn_teacher_profile") 
        self.btn_profile_detail.setCursor(Qt.PointingHandCursor)
        self.btn_profile_detail.clicked.connect(self.open_profile_dialog)

        self.btn_profile_detail.setStyleSheet("""
            QPushButton#btn_teacher_profile {
                background-color: transparent;
                color: #e67e22; 
                border: 1px solid #e67e22;
                border-radius: 15px;
                padding: 5px 15px;
                font-weight: bold;
                font-size: 17px;
                margin-top: 5px;
            }
            QPushButton#btn_teacher_profile:hover {
                background-color: #e67e22;
                color: white;
            }
        """)

        l.addWidget(self.avatar, alignment=Qt.AlignCenter)
        l.addWidget(self.lbl_tname)
        l.addWidget(self.lbl_tid)
        l.addWidget(self.btn_profile_detail, alignment=Qt.AlignCenter) 
        
        layout.addWidget(widget)


    def open_profile_dialog(self):
        if self.teacher_id:
            dialog = TeacherProfileDialog(self.teacher_id, self)
            dialog.exec_()

    def set_teacher_info(self, fullname, code):
        if fullname: self.lbl_tname.setText(fullname)
        if code: self.lbl_tid.setText(f"ID: {code}")
        
        possible_paths = [
            os.path.join(IMG_DIR, "teacher", f"{code}.jpg"),         
            os.path.join(IMG_DIR, "teacher", f"{code}.png"),         
            os.path.join(IMG_DIR, "teacher", f"{code.lower()}.jpg"),
            os.path.join(IMG_DIR, "teacher", f"{code.lower()}.png"), 
            os.path.join(IMG_DIR, "teacher", f"{fullname}.jpg")      
        ]

        final_path = None
        for p in possible_paths:
            if os.path.exists(p):
                final_path = p
                break
        
        if not final_path:
            final_path = os.path.join(IMG_DIR, "teacher_avatar.png")

        if os.path.exists(final_path):
            self.avatar.setText("") 

            src_pixmap = QPixmap(final_path)       
            if not src_pixmap.isNull():
                size = 80 
                
                rounded = QPixmap(size, size)
                rounded.fill(Qt.transparent)
                
                painter = QPainter(rounded)
                painter.setRenderHint(QPainter.Antialiasing, True) 
                painter.setRenderHint(QPainter.SmoothPixmapTransform, True)

                path = QPainterPath()
                path.addEllipse(0, 0, size, size)
                painter.setClipPath(path)
                painter.drawPixmap(0, 0, src_pixmap.scaled(size, size, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation))

                painter.setClipping(False) 

                pen = QPen(QColor("white"))
                pen.setWidth(3)
                painter.setPen(pen)
                painter.setBrush(Qt.NoBrush) 

                painter.drawEllipse(1, 1, size-2, size-2)
                
                painter.end()

                self.avatar.setStyleSheet("background: transparent;") 
                self.avatar.setPixmap(rounded)
            else:
                self.avatar.setText(code[:2])

    def create_stat_card(self, title, val, color):
        c = QFrame()
        c.setObjectName("stat_card")
        c.setFixedHeight(100)
        
        shadow = QGraphicsDropShadowEffect()
        shadow.setBlurRadius(15); shadow.setColor(QColor(0,0,0,30)); shadow.setYOffset(4)
        c.setGraphicsEffect(shadow)
        
        l = QVBoxLayout()
        l.setContentsMargins(25,15,25,15)
        
        t = QLabel(title)
        t.setStyleSheet("color: #7f8c8d; font-size: 13px; font-weight: bold; text-transform: uppercase;")
        
        v = QLabel(val)
        v.setObjectName("val")
        v.setStyleSheet(f"color: {color}; font-size: 32px; font-weight: 800;")
        
        c.setStyleSheet(f"#stat_card {{ border-left: 5px solid {color}; background: white; border-radius: 10px; }}")

        l.addWidget(t)
        l.addWidget(v)
        c.setLayout(l)
        return c

    def load_sidebar_data(self):
        if not self.teacher_id: return
        courses = get_teacher_courses(self.teacher_id)
        self.course_btns = []
        for i in reversed(range(self.courses_container.count())): 
            self.courses_container.itemAt(i).widget().setParent(None)
        
        for c in courses:
            b = QPushButton(c.course_name)
            b.setCheckable(True); b.setObjectName("sidebar_btn")
            b.clicked.connect(lambda _, cid=c.id, cn=c.course_name, cc=c.course_code, cr=c.credits: self.select_course(cid, cn, cc, cr))
            self.courses_container.addWidget(b); self.course_btns.append(b)
            
        if self.course_btns: self.course_btns[0].click()

    def select_course(self, cid, cname, ccode, credits):
        self.current_course_id = cid
        self.lbl_class.setText(cname)

        info_html = (
            f"<span style='color: #7f8c8d;'>Mã HP: </span>"
            f"<span style='color: #2980b9; font-weight: bold;'>{ccode}</span>"
            f"<span style='color: #bdc3c7;'>  |  </span>"
            f"<span style='color: #7f8c8d;'>Số tín chỉ: </span>"
            f"<span style='color: #e67e22; font-weight: bold;'>{credits}</span>"
        )
        self.lbl_sub_info.setText(info_html)
        
        for b in self.course_btns: b.setChecked(False)
        self.sender().setChecked(True)
        self.refresh_table()

    def refresh_table(self):
        if not self.current_course_id: return
        date = self.date_edit.date().toString("yyyy-MM-dd")
        data = get_attendance_list(self.current_course_id, date)

        if data: 
            data.sort(key=lambda x: (x['name'].strip().split()[-1], " ".join(x['name'].strip().split()[:-1])))
        
        self.all_table_data = data 
        

        self.table.setRowCount(len(data))
        for r, s in enumerate(data):
            self.table.setRowHeight(r, 60)
            
            # Cột 0: STT
            i_stt = QTableWidgetItem(str(r + 1))
            i_stt.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(r, 0, i_stt)
            
            # Cột 1: MSSV 
            i_mssv = QTableWidgetItem(s['mssv'])
            i_mssv.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(r, 1, i_mssv)
            
            # [XỬ LÝ TÁCH TÊN]
            full_name = s['name'].strip()
            parts = full_name.split()
            if len(parts) > 1:
                first_name = parts[-1]              
                last_name = " ".join(parts[:-1])   
            else:
                first_name = full_name
                last_name = ""
            # Cột 2: Họ đệm
            i_lastname = QTableWidgetItem(last_name)
            i_lastname.setTextAlignment(Qt.AlignCenter) 
            self.table.setItem(r, 2, i_lastname)

            # Cột 3: Tên (In đậm cho nổi bật)
            i_firstname = QTableWidgetItem(first_name)
            i_firstname.setFont(QFont("Segoe UI", 13)) 
            i_firstname.setTextAlignment(Qt.AlignCenter) 
            self.table.setItem(r, 3, i_firstname)

            # Cột 4: Ngày sinh
            raw_dob = s.get('dob', '')
            dob_display = str(raw_dob)
            if raw_dob and raw_dob != 'None':
                try: 
                    qdate = QDate.fromString(str(raw_dob), "yyyy-MM-dd")
                    if qdate.isValid(): dob_display = qdate.toString("dd-MM-yyyy")
                except: pass
            
            i_dob = QTableWidgetItem(dob_display)
            i_dob.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(r, 4, i_dob)

            # Cột 5: Lớp
            i_class = QTableWidgetItem(s.get('class_name', ''))
            i_class.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(r, 5, i_class)
            
            # Cột 6: Trạng thái
            w = AttendanceStatusBox(s['status'])
            w.currentTextChanged.connect(lambda t, sid=s['id']: self.handle_change(sid, t))
            self.table.setCellWidget(r, 6, w) 
            
            # Cột 7: Thời gian
            i_time = QTableWidgetItem(s['time'])
            i_time.setTextAlignment(Qt.AlignCenter)
            self.table.setItem(r, 7, i_time)
            
            # Cột 8: Ghi chú
            i_note = QTableWidgetItem(s['note'])
            i_note.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
            self.table.setItem(r, 8, i_note)
            
        self.filter_rows()
        self.refresh_stats_only()

    def handle_import_excel(self):
        if not self.teacher_id:
            QMessageBox.warning(self, "Lỗi", "Chưa xác định giáo viên (Teacher ID Null)")
            return

        path, _ = QFileDialog.getOpenFileName(self, "Chọn file lớp học", "", "Excel/CSV Files (*.xlsx *.xls *.csv)")
        if not path: return

        try:
            if path.endswith('.csv'):
                df_meta = pd.read_csv(path, header=None, nrows=1)
            else:
                df_meta = pd.read_excel(path, header=None, nrows=1)
            
            row_meta = df_meta.iloc[0]
            course_name = "Môn học mới"
            course_code = "UNKNOWN"
            
            # Logic tìm tên môn
            text_row = " ".join([str(x) for x in row_meta if str(x) != 'nan'])
            if "Tên môn học:" in text_row:
                parts = text_row.split("Mã môn học:")
                if len(parts) > 0:
                    course_name = parts[0].replace("Tên môn học:", "").strip().rstrip(",")
                if len(parts) > 1:
                    course_code = parts[1].strip()
            
            if course_code == "UNKNOWN":
                 try:
                     course_name = str(row_meta[0]).split(":")[-1].strip()
                     course_code = str(row_meta[4]).split(":")[-1].strip()
                 except: pass

            reply = QMessageBox.question(self, "Xác nhận nhập", 
                                       f"Môn: {course_name}\nMã: {course_code}\n\nBạn có muốn nhập?",
                                       QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.No: return

            cid = create_or_get_course(course_code, course_name, self.teacher_id)

            if path.endswith('.csv'):
                df = pd.read_csv(path, header=4, dtype=str)
            else:
                df = pd.read_excel(path, header=4, dtype=str)

            count_success = 0
            errors = []

            for index, row in df.iterrows():
                try:
                    mssv = str(row.iloc[1]).strip()
                    if not mssv or mssv.lower() in ['nan', 'mã sv', '']: continue

                    ho_dem = str(row.iloc[2]).strip()
                    ten = str(row.iloc[3]).strip()
                    if ho_dem.lower() == 'nan': ho_dem = ""
                    if ten.lower() == 'nan': ten = ""
                    full_name = f"{ho_dem} {ten}".strip()

                    raw_dob = str(row.iloc[4]).strip()
                    class_name = str(row.iloc[5]).strip()
                    if class_name.lower() == 'nan': class_name = ""

                    dob_val = None
                    if raw_dob and raw_dob.lower() != 'nan':
                        try:
                            dob_val = pd.to_datetime(raw_dob, dayfirst=True).date()
                        except:
                            errors.append(f"Dòng {index+6}: {mssv} - Lỗi ngày sinh ({raw_dob})")
                            continue 

                    success, msg = import_student_from_excel(mssv, full_name, dob_val, class_name, cid)
                    
                    if success:
                        count_success += 1
                    else:
                        errors.append(f"Dòng {index+6}: {mssv} - {msg}")

                except Exception as e:
                    errors.append(f"Dòng {index+6}: Lỗi không xác định - {str(e)}")

            msg_result = f"Đã nhập thành công: {count_success} sinh viên.\n"
            if errors:
                msg_result += f"\nCÓ {len(errors)} DÒNG LỖI:\n" + "\n".join(errors[:10])
                if len(errors) > 10: msg_result += "\n... và các dòng khác."
                QMessageBox.warning(self, "Kết quả Import (Có lỗi)", msg_result)
            else:
                QMessageBox.information(self, "Thành công", msg_result)

            self.load_sidebar_data()
            
        except Exception as e:
            QMessageBox.critical(self, "Lỗi Import", f"Lỗi nghiêm trọng: {e}")

    def filter_rows(self):
        search_text = self.txt_search.text().lower().strip()
        filter_status = self.cbb_filter.currentText()

        for r in range(self.table.rowCount()):
            mssv = self.table.item(r, 1).text().lower()

            ho_dem = self.table.item(r, 2).text().lower()
            ten = self.table.item(r, 3).text().lower()
            full_name = f"{ho_dem} {ten}"

            status_widget = self.table.cellWidget(r, 6)
            status = status_widget.currentText() if status_widget else ""

            match_search = (search_text == "") or (search_text in mssv) or (search_text in full_name)
            match_status = (filter_status == "Tất cả") or (status == filter_status)

            self.table.setRowHidden(r, not (match_search and match_status))

    def handle_change(self, sid, status):
        if not self.current_course_id: return
        date = self.date_edit.date().toString("yyyy-MM-dd")
        update_attendance(sid, self.current_course_id, date, status)
        self.refresh_stats_only()

    def refresh_stats_only(self):
        total = self.table.rowCount() 
        present = 0
        absent = 0
        for r in range(total):
            w = self.table.cellWidget(r, 6)
            if w:
                st = w.currentText()
                if st == "Có mặt": present += 1
                elif st == "Vắng": absent += 1
            
        self.card_total.findChild(QLabel, "val").setText(str(total))
        self.card_present.findChild(QLabel, "val").setText(str(present))
        self.card_absent.findChild(QLabel, "val").setText(str(absent))

    def open_camera(self):
        if not self.current_course_id:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn lớp học trước!")
            return
        date = self.date_edit.date().toString("yyyy-MM-dd")
        dialog = CameraDialog(self.current_course_id, date, self)
        dialog.exec_() 
        self.refresh_table()

    def open_add_student(self):
        if not self.current_course_id:
            QMessageBox.warning(self, "Cảnh báo", "Vui lòng chọn lớp học trước!")
            return
        self.dlg = AddStudentDialog(self.current_course_id, self)
        self.dlg.show()

    def export_excel(self):
        if self.table.rowCount() == 0:
            QMessageBox.warning(self, "Thông báo", "Không có dữ liệu để xuất!")
            return

        path, _ = QFileDialog.getSaveFileName(self, "Lưu file Excel", "", "Excel Files (*.xlsx)")
        
        if not path: 
            return

        if not path.lower().endswith('.xlsx'):
            path += '.xlsx'


        try:
            data = []
            for row in range(self.table.rowCount()):
                def get_text(r, c):
                    item = self.table.item(r, c)
                    return item.text().strip() if item else ""

                stt = get_text(row, 0)
                mssv = get_text(row, 1)
                ho_dem = get_text(row, 2)
                ten = get_text(row, 3) # Tên
                dob = get_text(row, 4)
                class_name = get_text(row, 5)

                # Lấy trạng thái
                status_widget = self.table.cellWidget(row, 6)
                if status_widget:
                    if hasattr(status_widget, 'currentText'):
                        status = status_widget.currentText()
                    elif hasattr(status_widget, 'text'):
                        status = status_widget.text()
                    else:
                        status = ""
                else:
                    status = ""
                
                time = get_text(row, 7)
                note = get_text(row, 8)

                data.append({
                    "STT": stt,
                    "MSSV": mssv,
                    "Họ đệm": ho_dem,
                    "Tên": ten,
                    "Ngày sinh": dob,
                    "Lớp hành chính": class_name,
                    "Trạng thái": status,
                    "Thời gian": time,
                    "Ghi chú": note
                })

            df = pd.DataFrame(data)

            course_name = self.lbl_class.text() 
            date_str = self.date_edit.date().toString("dd/MM/yyyy")
            report_title = f"BÁO CÁO ĐIỂM DANH: {course_name.upper()}"
            sub_title = f"Ngày điểm danh: {date_str}"
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='DiemDanh', index=False, startrow=3)
        
                workbook = writer.book
                worksheet = writer.sheets['DiemDanh']


                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                center_align = Alignment(horizontal='center', vertical='center')
                

                worksheet.merge_cells('A1:I1')
                cell_title = worksheet['A1']
                cell_title.value = report_title
                cell_title.font = Font(name='Arial', size=16, bold=True, color="000000")
                cell_title.alignment = center_align


                worksheet.merge_cells('A2:I2')
                cell_sub = worksheet['A2']
                cell_sub.value = sub_title
                cell_sub.font = Font(name='Arial', size=12, italic=True)
                cell_sub.alignment = center_align


                header_row = 4 
                for col_num, column_title in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=header_row, column=col_num)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.alignment = center_align
                    cell.border = thin_border


                data_len = len(data)
                for row in range(header_row + 1, header_row + 1 + data_len):
                    for col in range(1, 10):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
                        # Căn giữa các cột ngắn, căn trái cột Tên/Ghi chú
                        if col in [1, 2, 5, 7, 8]: 
                            cell.alignment = center_align
                        else:
                            cell.alignment = Alignment(horizontal='left', vertical='center')


                column_widths = [5, 15, 20, 10, 15, 15, 15, 15, 20]
                for i, width in enumerate(column_widths, 1):
                    worksheet.column_dimensions[chr(64 + i)].width = width

            QMessageBox.information(self, "Thành công", f"Đã xuất file thành công tại:\n{path}")

        except Exception as e:
            print(f"Error: {e}")
            QMessageBox.critical(self, "Lỗi", f"Có lỗi xảy ra khi xuất file:\n{str(e)}")